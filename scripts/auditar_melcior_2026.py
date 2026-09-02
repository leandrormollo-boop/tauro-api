#!/usr/bin/env python3
"""Auditoría read-only de la copia XLSX de MELCIOR 2026.

No modifica la planilla ni la base de datos. Resume encabezados reales,
importes, duplicados y excepciones antes de preparar una migración financiera.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel


MESES = (
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
)


def texto(valor: object) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def clave(valor: object) -> str:
    limpio = unicodedata.normalize("NFKD", texto(valor)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", "_", limpio.upper()).strip("_")


def dinero(valor: object) -> Decimal | None:
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float, Decimal)):
        if isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)):
            return None
        return Decimal(str(valor)).quantize(Decimal("0.01"))
    bruto = texto(valor).replace("$", "").replace("\u00a0", "").replace(" ", "")
    if not bruto or bruto in {"-", "—"} or bruto.startswith("="):
        return None
    if "," in bruto and "." in bruto:
        if bruto.rfind(",") > bruto.rfind("."):
            bruto = bruto.replace(".", "").replace(",", ".")
        else:
            bruto = bruto.replace(",", "")
    elif "," in bruto:
        bruto = bruto.replace(",", ".")
    try:
        return Decimal(bruto).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def tracking(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (int, float)) and float(valor).is_integer():
        return str(int(valor))
    return re.sub(r"\s+", "", texto(valor)).removesuffix(".0")


def fecha(valor: object, *, mes: int) -> str:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    if isinstance(valor, (int, float)) and 30_000 <= float(valor) <= 80_000:
        return from_excel(float(valor)).date().isoformat()
    s = texto(valor)
    if not s:
        return ""
    for patron in (r"^(\d{1,2})/(\d{1,2})/(\d{4})$", r"^(\d{4})-(\d{2})-(\d{2})"):
        m = re.match(patron, s)
        if not m:
            continue
        if patron.startswith("^(\\d{1,2})"):
            d, mo, y = map(int, m.groups())
        else:
            y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return ""
    if re.fullmatch(r"\d{1,2}", s):
        try:
            return date(2026, mes, int(s)).isoformat()
        except ValueError:
            return ""
    return ""


def ubicar_encabezado(ws) -> tuple[int, dict[str, int]]:
    if ws.max_row is None or ws.max_column is None:
        ws.calculate_dimension(force=True)
    for fila in range(1, min(ws.max_row, 15) + 1):
        mapa = {clave(ws.cell(fila, col).value): col for col in range(1, ws.max_column + 1)}
        if any(k.startswith("TRACKING") for k in mapa) and any(k.startswith("FACTURADO") for k in mapa):
            return fila, mapa
    raise ValueError(f"No se encontró encabezado operativo en {ws.title}")


def columna(mapa: dict[str, int], *prefijos: str) -> int | None:
    for prefijo in prefijos:
        for nombre, col in mapa.items():
            if nombre.startswith(prefijo):
                return col
    return None


def valor(ws, fila: int, col: int | None):
    return ws.cell(fila, col).value if col else None


def auditar(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    resultado: dict[str, object] = {"archivo": path.name, "meses": {}, "excepciones": []}
    todos: list[dict[str, object]] = []

    for numero_mes, nombre in enumerate(MESES, start=1):
        if nombre not in wb.sheetnames:
            resultado["meses"][nombre] = {"ausente": True}
            continue
        ws = wb[nombre]
        try:
            encabezado, mapa = ubicar_encabezado(ws)
        except ValueError:
            resultado["meses"][nombre] = {"sin_encabezado_operativo": True}
            continue
        c_fecha = columna(mapa, "FECHA") or columna(mapa, "MES")
        c_remitente = columna(mapa, "REMITENTE")
        c_destino = columna(mapa, "DESTINATARIO")
        c_pais = columna(mapa, "PAIS")
        c_peso = columna(mapa, "PESO")
        c_medidas = columna(mapa, "MEDIDAS")
        c_tipo = columna(mapa, "FLETE_O_TAX", "TIPO")
        c_tracking = columna(mapa, "TRACKING")
        c_facturado = columna(mapa, "FACTURADO")
        c_diferencia = columna(mapa, "DIF_INICIAL", "DIFERENCIAS", "DIFERENCIA")
        c_fc = columna(mapa, "NRO_FC", "NRO_DE_FC", "FC")
        c_estado = columna(mapa, "ESTADO")
        filas: list[dict[str, object]] = []
        vacias_seguidas = 0
        for fila in range(encabezado + 1, ws.max_row + 1):
            trk = tracking(valor(ws, fila, c_tracking))
            dest = texto(valor(ws, fila, c_destino))
            fact = dinero(valor(ws, fila, c_facturado))
            dif = dinero(valor(ws, fila, c_diferencia))
            tipo = clave(valor(ws, fila, c_tipo))
            if not any((trk, dest, fact not in (None, Decimal("0.00")), dif not in (None, Decimal("0.00")), tipo)):
                vacias_seguidas += 1
                if vacias_seguidas >= 100:
                    break
                continue
            vacias_seguidas = 0
            if tipo not in {"FLETE", "TAX"} and not trk:
                continue
            registro = {
                "mes": nombre,
                "fila": fila,
                "fecha": fecha(valor(ws, fila, c_fecha), mes=numero_mes),
                "remitente": texto(valor(ws, fila, c_remitente)),
                "destinatario": dest,
                "pais": texto(valor(ws, fila, c_pais)).upper(),
                "peso": texto(valor(ws, fila, c_peso)),
                "medidas": texto(valor(ws, fila, c_medidas)),
                "tipo": tipo or "FLETE",
                "tracking": trk,
                "facturado": str(fact) if fact is not None else "",
                "diferencia": str(dif) if dif is not None else "",
                "fc": tracking(valor(ws, fila, c_fc)),
                "estado": texto(valor(ws, fila, c_estado)).upper(),
            }
            filas.append(registro)
            todos.append(registro)

        conteo_tipo = Counter(r["tipo"] for r in filas)
        conteo_estado = Counter(r["estado"] or "SIN_ESTADO" for r in filas)
        faltantes = Counter()
        for r in filas:
            for campo in ("fecha", "tracking", "destinatario", "pais", "facturado"):
                if not r[campo]:
                    faltantes[campo] += 1
        resultado["meses"][nombre] = {
            "encabezado_fila": encabezado,
            "encabezados": {k: v for k, v in mapa.items() if k},
            "filas": len(filas),
            "tipos": dict(conteo_tipo),
            "estados": dict(conteo_estado),
            "facturado_ars": str(sum((Decimal(r["facturado"]) for r in filas if r["facturado"]), Decimal("0"))),
            "diferencias_ars": str(sum((Decimal(r["diferencia"]) for r in filas if r["diferencia"]), Decimal("0"))),
            "faltantes": dict(faltantes),
            "primera_fecha": min((r["fecha"] for r in filas if r["fecha"]), default=""),
            "ultima_fecha": max((r["fecha"] for r in filas if r["fecha"]), default=""),
        }

    grupos = defaultdict(list)
    for r in todos:
        grupos[(r["tracking"], r["tipo"])].append(r)
    duplicados = {
        f"{trk}|{tipo}": [{"mes": r["mes"], "fila": r["fila"], "facturado": r["facturado"], "diferencia": r["diferencia"]} for r in filas]
        for (trk, tipo), filas in grupos.items() if trk and len(filas) > 1
    }
    resultado["resumen"] = {
        "filas": len(todos),
        "guias_unicas": len({r["tracking"] for r in todos if r["tracking"]}),
        "facturado_ars": str(sum((Decimal(r["facturado"]) for r in todos if r["facturado"]), Decimal("0"))),
        "diferencias_ars": str(sum((Decimal(r["diferencia"]) for r in todos if r["diferencia"]), Decimal("0"))),
        "tipos": dict(Counter(r["tipo"] for r in todos)),
        "duplicados_tracking_tipo": duplicados,
    }

    detalle = wb["DETALLE 2026"] if "DETALLE 2026" in wb.sheetnames else None
    if detalle:
        if detalle.max_row is None or detalle.max_column is None:
            detalle.calculate_dimension(force=True)
        pagos = []
        for fila in range(8, min(detalle.max_row, 100) + 1):
            monto = dinero(detalle.cell(fila, 1).value)
            descripcion = texto(detalle.cell(fila, 2).value)
            if monto is not None and monto != 0:
                pagos.append({"fila": fila, "monto_ars": str(monto), "detalle": descripcion})
        resultado["detalle_2026"] = {
            "saldo_pendiente_2025": str(dinero(detalle["A5"].value) or Decimal("0")),
            "pagos": pagos,
        }
    resultado["filas"] = todos
    return resultado


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--json", type=Path)
    args = parser.parse_args()
    resultado = auditar(args.xlsx)
    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")
    compacto = {k: v for k, v in resultado.items() if k != "filas"}
    print(json.dumps(compacto, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
