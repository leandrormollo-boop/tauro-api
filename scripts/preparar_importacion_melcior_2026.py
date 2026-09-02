#!/usr/bin/env python3
"""Prepara un lote idempotente y auditable para MELCIOR 2026.

La preparación es read-only. Cruza la hoja del cliente con TAURO 2026,
resuelve únicamente duplicados con evidencia inequívoca y deja el resto en
cuarentena. No escribe en PostgreSQL.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from auditar_melcior_2026 import auditar, clave, dinero, fecha, texto, tracking


def _fecha_maestra(valor: object) -> str:
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    return fecha(valor, mes=1)


def leer_maestra(path: Path) -> list[dict[str, object]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["ENVIOS 2026"]
    ws.calculate_dimension(force=True)
    encabezados = {clave(v): i for i, v in enumerate(next(ws.iter_rows(values_only=True)))}

    def col(nombre: str) -> int:
        return encabezados[nombre]

    filas = []
    for nro, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        remitente = texto(row[col("REMITENTE")]).upper()
        if remitente != "JUAN PABLO MELCIOR":
            continue
        fecha_iso = _fecha_maestra(row[col("FECHA")])
        if not fecha_iso.startswith("2026-"):
            continue
        tipo = clave(row[col("FLETE_O_TAX")]) or "FLETE"
        filas.append({
            "fila": nro,
            "empresa": texto(row[col("EMPRESA")]).upper(),
            "fecha": fecha_iso,
            "remitente": remitente,
            "destinatario": texto(row[col("DESTINATARIO")]),
            "pais": texto(row[col("PAIS")]).upper(),
            "peso": texto(row[col("PESO")]),
            "medidas": texto(row[col("MEDIDAS")]),
            "tracking": tracking(row[col("TRACKING")]),
            "tipo": tipo,
            "fc": tracking(row[col("NRO_FC")]),
            "facturado": str(dinero(row[col("FACTURADO")]) or ""),
            "diferencia": str(dinero(row[col("DIF_INICIAL_VS_FC")]) or ""),
            "saldo_ars": str(dinero(row[col("SALDO_ARS")]) or ""),
            "costo_inicial": str(dinero(row[col("COSTOINICIAL")]) or ""),
            "estado": texto(row[col("ESTADO")]).upper(),
        })
    return filas


def _decimal(valor: object) -> Decimal | None:
    return Decimal(str(valor)) if valor not in (None, "") else None


def seleccionar_cliente(
    cliente: list[dict[str, object]], maestra: list[dict[str, object]]
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    clientes_por_tracking = defaultdict(list)
    for fila in cliente:
        if fila["tipo"] == "FLETE" and fila["tracking"]:
            clientes_por_tracking[fila["tracking"]].append(fila)
    maestras_por_tracking = defaultdict(list)
    for fila in maestra:
        if fila["tipo"] == "FLETE" and fila["tracking"]:
            maestras_por_tracking[fila["tracking"]].append(fila)

    seleccionadas = []
    resueltas = []
    cuarentena = []
    for trk, candidatas in sorted(clientes_por_tracking.items()):
        maestras = maestras_por_tracking.get(trk, [])
        elegida = None
        razon = ""
        if len(candidatas) == 1:
            elegida = candidatas[0]
            razon = "TRACKING_UNICO_CLIENTE"
        else:
            importes_maestra = Counter(m["facturado"] for m in maestras if m["facturado"])
            exactas = [c for c in candidatas if c["facturado"] and importes_maestra[c["facturado"]] == 1]
            if len(maestras) == 1 and len(exactas) == 1:
                elegida = exactas[0]
                razon = "IMPORTE_EXACTO_UNICA_FILA_MAESTRA"
            else:
                maestras_fc = [m for m in maestras if m["fc"]]
                if len(maestras_fc) == 1:
                    con_fc = [c for c in candidatas if c["facturado"] == maestras_fc[0]["facturado"]]
                    if len(con_fc) == 1:
                        elegida = con_fc[0]
                        razon = "FILA_MAESTRA_CON_FC_UNICA"
        if elegida is None:
            cuarentena.append({"tracking": trk, "cliente": candidatas, "maestra": maestras})
            continue
        maestras_compatibles = [m for m in maestras if m["facturado"] == elegida["facturado"]]
        maestra = maestras_compatibles[0] if len(maestras_compatibles) == 1 else (
            maestras[0] if len(maestras) == 1 else None
        )
        combinada = dict(elegida)
        combinada["maestra"] = maestra
        combinada["resolucion"] = razon
        seleccionadas.append(combinada)
        if len(candidatas) > 1:
            resueltas.append({
                "tracking": trk,
                "razon": razon,
                "fila_elegida": {"mes": elegida["mes"], "fila": elegida["fila"]},
                "filas_descartadas": [
                    {"mes": c["mes"], "fila": c["fila"]}
                    for c in candidatas if c is not elegida
                ],
            })
    return seleccionadas, resueltas, cuarentena


def preparar(cliente_path: Path, maestra_path: Path) -> dict[str, object]:
    auditoria_cliente = auditar(cliente_path)
    cliente = auditoria_cliente["filas"]
    maestra = leer_maestra(maestra_path)
    seleccionadas, resueltas, cuarentena = seleccionar_cliente(cliente, maestra)

    con_diferencia = [r for r in seleccionadas if _decimal(r["diferencia"]) not in (None, Decimal("0"))]
    evidencia_costos = []
    sin_evidencia_costos = []
    diferencias_no_reconciliadas = []
    for fila in con_diferencia:
        m = fila.get("maestra") or {}
        costo_inicial = _decimal(m.get("costo_inicial"))
        costo_real = _decimal(m.get("saldo_ars"))
        diferencia = _decimal(fila["diferencia"])
        if costo_inicial is None or costo_real is None:
            sin_evidencia_costos.append({"tracking": fila["tracking"], "mes": fila["mes"], "fila": fila["fila"]})
            continue
        evidencia_costos.append(fila["tracking"])
        if abs((costo_real - costo_inicial) - diferencia) > Decimal("0.02"):
            diferencias_no_reconciliadas.append({
                "tracking": fila["tracking"],
                "diferencia_cliente": str(diferencia),
                "saldo_menos_costo_inicial": str(costo_real - costo_inicial),
                "fila_cliente": f"{fila['mes']}!{fila['fila']}",
                "fila_maestra": m.get("fila"),
            })

    facturado = sum((_decimal(r["facturado"]) or Decimal("0") for r in seleccionadas), Decimal("0"))
    diferencias = sum((_decimal(r["diferencia"]) or Decimal("0") for r in seleccionadas), Decimal("0"))
    pagos = auditoria_cliente.get("detalle_2026", {}).get("pagos", [])
    pagos_total = sum((_decimal(p["monto_ars"]) or Decimal("0") for p in pagos), Decimal("0"))
    saldo_2025 = _decimal(auditoria_cliente.get("detalle_2026", {}).get("saldo_pendiente_2025")) or Decimal("0")
    saldo = facturado + diferencias + saldo_2025 - pagos_total
    huella = hashlib.sha256(cliente_path.read_bytes() + maestra_path.read_bytes()).hexdigest()
    return {
        "source_sha256": huella,
        "resumen": {
            "filas_cliente_crudas": len(cliente),
            "envios_flete_seleccionados": len(seleccionadas),
            "duplicados_resueltos": len(resueltas),
            "duplicados_en_cuarentena": len(cuarentena),
            "envios_con_diferencia": len(con_diferencia),
            "diferencias_con_costos_maestros": len(evidencia_costos),
            "diferencias_sin_costos_maestros": len(sin_evidencia_costos),
            "diferencias_no_reconciliadas": len(diferencias_no_reconciliadas),
            "facturado_inicial_ars": str(facturado),
            "diferencias_ars": str(diferencias),
            "saldo_pendiente_2025_ars": str(saldo_2025),
            "pagos_ars": str(pagos_total),
            "saldo_resultante_ars": str(saldo),
        },
        "duplicados_resueltos": resueltas,
        "cuarentena": cuarentena,
        "diferencias_sin_costos_maestros": sin_evidencia_costos,
        "diferencias_no_reconciliadas": diferencias_no_reconciliadas,
        "pagos_sin_fecha": pagos,
        "envios": seleccionadas,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--cliente", type=Path, required=True)
    parser.add_argument("--maestra", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    lote = preparar(args.cliente, args.maestra)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(lote, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({k: v for k, v in lote.items() if k in {"source_sha256", "resumen", "duplicados_resueltos", "cuarentena", "diferencias_sin_costos_maestros", "diferencias_no_reconciliadas", "pagos_sin_fecha"}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
