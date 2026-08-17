# ============================================================
# Backup del cliente en Excel — SUS datos, para SU PC.
# ============================================================
# Pedido de Leandro (28/07): cada cliente puede bajarse desde el portal
# un .xlsx con sus envíos, su cuenta corriente y su catálogo, para
# controlarlo en su propia planilla. No confundir con /admin/backup.json,
# que es la copia de TODA la plataforma para recuperación de desastres.
#
# Sólo datos del cliente autenticado: nada de otros clientes, nada de
# secretos, nada de tokens.
# ============================================================
from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Iterable, Iterator, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from servicios.couriers_urls import ambito_envio

# Identidad de marca en el archivo: violeta TAURO, no colores default.
_VIOLETA = "5B3AD4"
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor=_VIOLETA)


def _hoja(
    wb: Workbook,
    titulo: str,
    columnas: list[str],
    filas: Iterable[Sequence],
):
    """Crea una hoja consumiendo las filas una sola vez.

    Aceptar iterables permite exportar cuentas extensas página por página sin
    construir otra copia completa del historial en memoria.
    """
    ws = wb.create_sheet(titulo)
    ws.append(columnas)
    for c in range(1, len(columnas) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = _HEADER_FONT
        cel.fill = _HEADER_FILL
        cel.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    anchos = [len(str(columna)) for columna in columnas]
    for numero, fila in enumerate(filas, start=1):
        fila = list(fila)
        ws.append(fila)
        # Medir sólo una muestra acotada; el contenido igualmente se escribe
        # completo. Una observación enorme no puede deformar toda la hoja.
        if numero <= 200:
            for indice, valor in enumerate(fila[:len(columnas)]):
                anchos[indice] = max(anchos[indice], len(str(valor or "")))

    for c, mejor in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(c)].width = min(mejor + 3, 42)

    ws.auto_filter.ref = ws.dimensions
    return ws


_COLUMNAS_CUENTA = [
    "Fecha", "Tipo", "Ámbito", "Concepto", "Referencia", "Estado",
    "Debe (ARS)", "Haber (ARS)", "Documento",
]
_TIPOS_CUENTA = {
    "FC": "Factura",
    "PENDIENTE_FACTURA": "Pendiente de facturación",
    "PAGO": "Pago aplicado",
    "PAGO_PENDIENTE": "Pago en revisión",
}
_CUENTA_PAGE_SIZE = 100
_FORMATO_DINERO = '#,##0.00'


def _iter_movimientos_cuenta(
    cliente_id: str,
    ambito: str,
    *,
    solo_ambito: Optional[str] = None,
) -> Iterator[dict]:
    """Recorre todo un ledger con paginación numerada y avance verificado."""
    from servicios.cuenta_corriente import movimientos_cuenta_paginados

    pagina = 1
    while True:
        resultado = movimientos_cuenta_paginados(
            cliente_id, ambito, "todos", pagina, _CUENTA_PAGE_SIZE
        )
        for movimiento in resultado.get("items") or []:
            if solo_ambito is None or movimiento.get("ambito") == solo_ambito:
                yield movimiento

        actual = int(resultado.get("pagina_actual") or pagina)
        total_paginas = max(1, int(resultado.get("total_paginas") or 1))
        if actual >= total_paginas:
            break
        if actual < pagina:
            raise RuntimeError("La paginación de la cuenta retrocedió durante el export.")
        pagina = actual + 1


def _fila_movimiento_cuenta(movimiento: dict) -> list:
    """Una fila contable sin conversiones binarias de dinero."""
    return [
        movimiento.get("fecha") or "",
        _TIPOS_CUENTA.get(movimiento.get("tipo"), movimiento.get("tipo") or ""),
        str(movimiento.get("ambito") or "").replace("_", " ").title(),
        movimiento.get("concepto") or "",
        movimiento.get("referencia") or "",
        movimiento.get("estado") or "",
        Decimal(str(movimiento.get("debe_ars") or 0)),
        Decimal(str(movimiento.get("haber_ars") or 0)),
        movimiento.get("archivo_url") or "",
    ]


def _hoja_cuenta(
    wb: Workbook,
    titulo: str,
    cliente_id: str,
    ambito: str,
    ledger: dict,
    *,
    solo_ambito: Optional[str] = None,
) -> int:
    """Escribe movimientos paginados y un cierre basado en el resumen oficial."""
    cantidad = 0
    total_debe = Decimal("0.00")
    total_haber = Decimal("0.00")

    def filas():
        nonlocal cantidad, total_debe, total_haber
        for movimiento in _iter_movimientos_cuenta(
            cliente_id, ambito, solo_ambito=solo_ambito
        ):
            cantidad += 1
            fila = _fila_movimiento_cuenta(movimiento)
            total_debe += fila[6]
            total_haber += fila[7]
            yield fila
        yield []
        yield ["", "", "", "", "", "FACTURADO", ledger["facturado_ars"]]
        yield ["", "", "", "", "", "PENDIENTE DE FACTURACIÓN", ledger["pendiente_facturacion_ars"]]
        yield ["", "", "", "", "", "TOTAL DEBE", ledger["debe_ars"]]
        yield ["", "", "", "", "", "TOTAL HABER", "", ledger["haber_ars"]]
        yield ["", "", "", "", "", "SALDO (DEBE - HABER)", ledger["saldo_ars"]]

    ws = _hoja(wb, titulo, _COLUMNAS_CUENTA, filas())
    if total_debe != ledger["debe_ars"] or total_haber != ledger["haber_ars"]:
        raise RuntimeError(
            f"{titulo} no reconcilia sus movimientos contra el resumen."
        )
    for fila in ws.iter_rows(min_row=2, min_col=7, max_col=8):
        for celda in fila:
            if celda.value not in (None, ""):
                celda.number_format = _FORMATO_DINERO
    return cantidad


def _validar_reconciliacion(resumen: dict) -> None:
    """Frena el backup si los libros por ámbito no cierran contra el total."""
    nacional = resumen["nacional"]
    internacional = resumen["internacional"]
    consolidado = resumen["consolidado"]
    sin_imputar = resumen["credito_sin_imputar_ars"]
    sin_clasificar = resumen["cargos_sin_clasificar_ars"]
    if consolidado["debe_ars"] != (
        nacional["debe_ars"] + internacional["debe_ars"] + sin_clasificar
    ):
        raise RuntimeError("La cuenta no reconcilia: el debe consolidado no cierra.")
    if consolidado["haber_ars"] != (
        nacional["haber_ars"] + internacional["haber_ars"] + sin_imputar
    ):
        raise RuntimeError("La cuenta no reconcilia: el haber consolidado no cierra.")
    if consolidado["saldo_ars"] != (
        nacional["saldo_ars"] + internacional["saldo_ars"]
        + sin_clasificar - sin_imputar
    ):
        raise RuntimeError("La cuenta no reconcilia: el saldo consolidado no cierra.")


def generar_excel_cliente(cliente_id: str) -> bytes:
    """El archivo completo, en memoria. Nunca toca disco."""
    from servicios.catalogo import get_productos
    from servicios.cuenta_corriente import resumen_cuenta_por_ambito
    from servicios.solicitudes_guia import listar_solicitudes_cliente

    cliente_id = cliente_id.strip().upper()
    wb = Workbook()
    wb.remove(wb.active)   # la hoja default vacía

    # ── Envíos ──────────────────────────────────────────────
    solicitudes = listar_solicitudes_cliente(cliente_id, limite=None)

    def _nac_o_int(s: dict) -> str:
        ambito = ambito_envio(s)
        return {
            "nacional": "Nacional",
            "internacional": "Internacional",
        }.get(ambito, "Sin clasificar")

    filas_por_ambito = {
        "nacional": [],
        "internacional": [],
        "sin_clasificar": [],
    }
    for s in solicitudes:
        creado = s.get("created_at")
        ambito = ambito_envio(s)
        fila = [
            creado.strftime("%d/%m/%Y") if creado else "",
            s.get("estado") or "",
            _nac_o_int(s),
            (s.get("courier") or "FEDEX").upper(),
            s.get("tracking") or "",
            s.get("producto_alias") or "",
            s.get("cantidad") or 1,
            s.get("dest_nombre") or "",
            s.get("dest_ciudad") or "",
            s.get("destino_pais") or "",
            float(s.get("peso_kg") or 0),
            float(s.get("valor_declarado_usd") or 0),
            float(s.get("precio_tauro_ars") or 0),
            float(s.get("precio_tauro_usd") or 0),
            s.get("observaciones") or "",
        ]
        filas_por_ambito.setdefault(ambito, filas_por_ambito["sin_clasificar"]).append(fila)

    columnas_envio = [
        "Fecha", "Estado", "Tipo", "Courier", "Tracking", "Producto", "Cajas",
        "Destinatario", "Ciudad", "País", "Peso (kg)", "Valor declarado (USD)",
        "Costo (ARS)", "Costo (USD)", "Observaciones",
    ]
    _hoja(wb, "Envios_Nacionales", columnas_envio, filas_por_ambito["nacional"])
    _hoja(wb, "Envios_Internacionales", columnas_envio, filas_por_ambito["internacional"])
    if filas_por_ambito["sin_clasificar"]:
        _hoja(wb, "Envios_Sin_clasificar", columnas_envio,
              filas_por_ambito["sin_clasificar"])

    # ── Cuenta corriente ────────────────────────────────────
    resumen_cuenta = resumen_cuenta_por_ambito(cliente_id)
    _validar_reconciliacion(resumen_cuenta)
    _hoja_cuenta(
        wb, "Cuenta_Consolidada", cliente_id, "consolidado",
        resumen_cuenta["consolidado"],
    )
    _hoja_cuenta(
        wb, "Cuenta_Nacional", cliente_id, "nacional",
        resumen_cuenta["nacional"],
    )
    _hoja_cuenta(
        wb, "Cuenta_Internacional", cliente_id, "internacional",
        resumen_cuenta["internacional"],
    )

    cero = Decimal("0.00")
    credito_sin_imputar = resumen_cuenta["credito_sin_imputar_ars"]
    ledger_sin_imputar = {
        "facturado_ars": cero,
        "pendiente_facturacion_ars": cero,
        "debe_ars": cero,
        "haber_ars": credito_sin_imputar,
        "saldo_ars": -credito_sin_imputar,
    }
    cantidad_sin_imputar = _hoja_cuenta(
        wb, "Cuenta_Sin_imputar", cliente_id, "consolidado",
        ledger_sin_imputar, solo_ambito="SIN_IMPUTAR",
    )
    if cantidad_sin_imputar == 0 and credito_sin_imputar == 0:
        wb.remove(wb["Cuenta_Sin_imputar"])

    consolidado = resumen_cuenta["consolidado"]
    nacional = resumen_cuenta["nacional"]
    internacional = resumen_cuenta["internacional"]
    cargos_sin_clasificar = resumen_cuenta["cargos_sin_clasificar_ars"]
    ledger_sin_clasificar = {
        "facturado_ars": (
            consolidado["facturado_ars"] - nacional["facturado_ars"]
            - internacional["facturado_ars"]
        ),
        "pendiente_facturacion_ars": (
            consolidado["pendiente_facturacion_ars"]
            - nacional["pendiente_facturacion_ars"]
            - internacional["pendiente_facturacion_ars"]
        ),
        "debe_ars": cargos_sin_clasificar,
        "haber_ars": cero,
        "saldo_ars": cargos_sin_clasificar,
    }
    cantidad_sin_clasificar = _hoja_cuenta(
        wb, "Cuenta_Sin_clasificar", cliente_id, "consolidado",
        ledger_sin_clasificar, solo_ambito="SIN_CLASIFICAR",
    )
    if cantidad_sin_clasificar == 0 and cargos_sin_clasificar == 0:
        wb.remove(wb["Cuenta_Sin_clasificar"])

    # ── Catálogo ────────────────────────────────────────────
    productos = get_productos(cliente_id, solo_activos=False)
    filas_prod = [[
        p.alias_interno, p.nombre_invoice, p.hs_code,
        p.largo_cm, p.ancho_cm, p.alto_cm, p.peso_kg,
        p.valor_usd_default, "Aprobado" if p.activo else "En revisión",
    ] for p in productos]
    _hoja(wb, "Mis productos", [
        "SKU", "Descripción invoice", "HS code", "Largo (cm)", "Ancho (cm)",
        "Alto (cm)", "Peso (kg)", "Valor (USD)", "Estado",
    ], filas_prod)

    # ── Portada mínima ──────────────────────────────────────
    ws = wb.create_sheet("Resumen", 0)
    ws["A1"] = f"TAURO Solutions — backup de {cliente_id}"
    ws["A1"].font = Font(bold=True, size=14, color=_VIOLETA)
    ws["A2"] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A4"] = "Envíos nacionales"
    ws["B4"] = len(filas_por_ambito["nacional"])
    ws["A5"] = "Envíos internacionales"
    ws["B5"] = len(filas_por_ambito["internacional"])
    ws["A6"] = "Históricos sin clasificar"
    ws["B6"] = len(filas_por_ambito["sin_clasificar"])
    ws["A7"] = "Debe consolidado (ARS)"
    ws["B7"] = consolidado["debe_ars"]
    ws["A8"] = "Haber consolidado (ARS)"
    ws["B8"] = consolidado["haber_ars"]
    ws["A9"] = "Saldo total consolidado (ARS)"
    ws["B9"] = consolidado["saldo_ars"]
    ws["A10"] = "Saldo nacional (ARS)"
    ws["B10"] = nacional["saldo_ars"]
    ws["A11"] = "Saldo internacional (ARS)"
    ws["B11"] = internacional["saldo_ars"]
    ws["A12"] = "Crédito sin imputar (ARS)"
    ws["B12"] = credito_sin_imputar
    ws["A13"] = "Cargos sin clasificar (ARS)"
    ws["B13"] = cargos_sin_clasificar
    ws["A14"] = "Productos en catálogo"
    ws["B14"] = len(filas_prod)
    for fila in range(7, 14):
        ws.cell(row=fila, column=2).number_format = _FORMATO_DINERO
    ws.column_dimensions["A"].width = 31

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
