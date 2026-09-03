"""Importación auditada de la base interna histórica de WAIMAO 2026.

La fuente válida es ``COSTOINICIAL`` de la hoja ``ENVIOS 2026`` del libro
madre. ``SALDO ARS`` no se usa: representa la factura real del courier y
pertenece a conciliación, no a la cotización aceptada.
"""

from __future__ import annotations

import hashlib
import math
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl

from core.database import get_conn
from servicios.conciliacion_couriers import (
    CENTAVO_CONTROL,
    ConciliacionCourierError,
    registrar_snapshot_cotizacion,
)


HOJA = "ENVIOS 2026"
CLIENTE_ID = "WAIMAO"
ORIGEN = "IMPORT_SHEET_2026"
CUATRO = Decimal("0.0001")


class ImportacionCostoWaimaoError(RuntimeError):
    """La evidencia no permite una importación financiera segura."""


def _texto(valor: Any) -> str:
    return str(valor or "").strip()


def _clave(valor: Any) -> str:
    if isinstance(valor, (int, Decimal)) and not isinstance(valor, bool):
        bruto = str(valor)
    elif isinstance(valor, float) and math.isfinite(valor) and valor.is_integer():
        # Google/Excel suele materializar trackings numéricos como 123.0.
        bruto = str(int(valor))
    else:
        bruto = _texto(valor)
        if re.fullmatch(r"\d+\.0+", bruto):
            bruto = bruto.split(".", 1)[0]
    return re.sub(r"[^A-Z0-9]", "", bruto.upper())


def _dinero(valor: Any) -> Decimal | None:
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)):
        return None
    bruto = _texto(valor).replace("$", "").replace("\u00a0", "").replace(" ", "")
    if not bruto or bruto in {"-", "—"} or bruto.startswith("="):
        return None
    if "," in bruto and "." in bruto:
        if bruto.rfind(",") > bruto.rfind("."):
            bruto = bruto.replace(".", "").replace(",", ".")
        else:
            bruto = bruto.replace(",", "")
    elif "," in bruto:
        bruto = bruto.replace(",", ".")
    try:
        numero = Decimal(bruto)
    except InvalidOperation as exc:
        raise ImportacionCostoWaimaoError(
            f"Importe inválido en la planilla: {valor!r}."
        ) from exc
    if not numero.is_finite():
        raise ImportacionCostoWaimaoError("La planilla contiene un importe no finito.")
    return numero.quantize(CUATRO, rounding=ROUND_HALF_UP)


def leer_costos_iniciales_waimao(ruta_xlsx: str | Path) -> dict[str, Any]:
    """Lee evidencia sin modificar el XLSX y devuelve una fila por tracking."""
    ruta = Path(ruta_xlsx)
    source_sha256 = hashlib.sha256(ruta.read_bytes()).hexdigest()
    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    try:
        if HOJA not in libro.sheetnames:
            raise ImportacionCostoWaimaoError(f"Falta la hoja {HOJA!r}.")
        hoja = libro[HOJA]
        encabezados = {
            _texto(celda.value).upper(): indice
            for indice, celda in enumerate(next(hoja.iter_rows(min_row=1, max_row=1)))
            if _texto(celda.value)
        }
        requeridos = {
            "EMPRESA", "REMITENTE", "TRACKING", "FLETE O TAX",
            "FACTURADO", "SALDO ARS", "COSTOINICIAL",
        }
        faltantes = sorted(requeridos - encabezados.keys())
        if faltantes:
            raise ImportacionCostoWaimaoError(
                "Faltan encabezados obligatorios: " + ", ".join(faltantes)
            )

        evidencias: dict[str, dict[str, Any]] = {}
        omitidas: list[dict[str, Any]] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), 2):
            def valor(campo: str) -> Any:
                indice = encabezados[campo]
                return fila[indice] if indice < len(fila) else None

            if _texto(valor("REMITENTE")).upper() != CLIENTE_ID:
                continue
            tracking = _clave(valor("TRACKING"))
            concepto = _texto(valor("FLETE O TAX")).upper()
            costo = _dinero(valor("COSTOINICIAL"))
            if concepto != "FLETE" or not tracking or costo is None or costo <= 0:
                omitidas.append({
                    "fila_sheet": numero_fila,
                    "tracking": tracking or None,
                    "motivo": (
                        "NO_ES_FLETE" if concepto and concepto != "FLETE"
                        else "SIN_TRACKING" if not tracking
                        else "SIN_COSTO_INICIAL"
                    ),
                })
                continue
            evidencia = {
                "fila_sheet": numero_fila,
                "tracking": tracking,
                "empresa_sheet": _texto(valor("EMPRESA")).upper(),
                "costo_inicial_ars": costo,
                "facturado_sheet_ars": _dinero(valor("FACTURADO")),
            }
            anterior = evidencias.get(tracking)
            if anterior and anterior["costo_inicial_ars"] != costo:
                raise ImportacionCostoWaimaoError(
                    f"El tracking {tracking} tiene costos iniciales incompatibles."
                )
            evidencias[tracking] = evidencia
        return {
            "ruta": str(ruta),
            "source_sha256": source_sha256,
            "evidencias": list(evidencias.values()),
            "omitidas": omitidas,
        }
    finally:
        libro.close()


def _cargar_envios_waimao() -> list[dict[str, Any]]:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT s.id AS solicitud_id, s.coti_id, s.courier,
                       s.servicio_courier, s.tracking, s.precio_tauro_ars,
                       s.peso_kg, s.bultos, s.created_at,
                       s.estado AS solicitud_estado, COALESCE(s.test, FALSE) AS test,
                       e.id AS envio_id, e.estado AS envio_estado,
                       snap.id AS snapshot_id,
                       snap.costo_courier_estimado_ars AS snapshot_costo_ars,
                       snap.precio_cliente_inicial_ars AS snapshot_precio_ars
                  FROM solicitudes_guia s
             LEFT JOIN envios e ON e.solicitud_id = s.id
             LEFT JOIN envio_cotizacion_snapshots snap ON snap.solicitud_id = s.id
                 WHERE UPPER(BTRIM(s.cliente_id)) = %s
                 ORDER BY s.id
                """,
                (CLIENTE_ID,),
            )
            return [dict(fila) for fila in cur.fetchall()]


def planificar_importacion_costos_waimao(
    ruta_xlsx: str | Path,
) -> dict[str, Any]:
    """Cruza planilla/base y falla ante cualquier ambigüedad financiera."""
    fuente = leer_costos_iniciales_waimao(ruta_xlsx)
    envios = _cargar_envios_waimao()
    por_tracking: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for envio in envios:
        tracking = _clave(envio.get("tracking"))
        if tracking:
            por_tracking[tracking].append(envio)

    candidatos: list[dict[str, Any]] = []
    existentes: list[dict[str, Any]] = []
    no_aplicables: list[dict[str, Any]] = []
    evidencia_por_tracking = {
        fila["tracking"]: fila for fila in fuente["evidencias"]
    }

    for evidencia in fuente["evidencias"]:
        tracking = evidencia["tracking"]
        coincidencias = por_tracking.get(tracking, [])
        if not coincidencias:
            no_aplicables.append({"tracking": tracking, "motivo": "NO_EXISTE_EN_BASE"})
            continue
        if len(coincidencias) != 1:
            raise ImportacionCostoWaimaoError(
                f"El tracking {tracking} coincide con {len(coincidencias)} solicitudes."
            )
        envio = coincidencias[0]
        if envio.get("test") or envio.get("solicitud_estado") == "CANCELADO":
            no_aplicables.append({"tracking": tracking, "motivo": "NO_OPERATIVO"})
            continue
        if not envio.get("envio_id") or envio.get("envio_estado") != "ACTIVO":
            raise ImportacionCostoWaimaoError(
                f"El tracking {tracking} no tiene un cargo ACTIVO único."
            )
        try:
            precio = Decimal(str(envio["precio_tauro_ars"])).quantize(CUATRO)
        except (InvalidOperation, TypeError, ValueError) as exc:
            raise ImportacionCostoWaimaoError(
                f"El tracking {tracking} no tiene precio aceptado válido."
            ) from exc
        costo = evidencia["costo_inicial_ars"]
        if costo > precio + CENTAVO_CONTROL:
            raise ImportacionCostoWaimaoError(
                f"El costo inicial de {tracking} supera el precio aceptado."
            )
        item = {
            **evidencia,
            **envio,
            "precio_tauro_ars": precio,
            "margen_tauro_ars": (precio - costo).quantize(CUATRO),
        }
        if envio.get("snapshot_id"):
            costo_existente = Decimal(str(envio["snapshot_costo_ars"])).quantize(CUATRO)
            precio_existente = Decimal(str(envio["snapshot_precio_ars"])).quantize(CUATRO)
            if (
                abs(costo_existente - costo) > CENTAVO_CONTROL
                or abs(precio_existente - precio) > CENTAVO_CONTROL
            ):
                raise ImportacionCostoWaimaoError(
                    f"El snapshot existente de {tracking} contradice la planilla."
                )
            existentes.append(item)
        else:
            candidatos.append(item)

    sin_evidencia = []
    for envio in envios:
        tracking = _clave(envio.get("tracking"))
        if (
            tracking and not envio.get("test")
            and envio.get("solicitud_estado") != "CANCELADO"
            and envio.get("envio_estado") == "ACTIVO"
            and not envio.get("snapshot_id")
            and tracking not in evidencia_por_tracking
        ):
            sin_evidencia.append({
                "solicitud_id": int(envio["solicitud_id"]),
                "tracking": tracking,
                "motivo": "SIN_COSTOINICIAL_EN_SHEET",
            })

    return {
        "source_sha256": fuente["source_sha256"],
        "candidatos": candidatos,
        "existentes": existentes,
        "no_aplicables": no_aplicables,
        "sin_evidencia": sin_evidencia,
        "omitidas_sheet": fuente["omitidas"],
    }


def importar_costos_waimao(
    ruta_xlsx: str | Path,
    *,
    actor: str,
) -> dict[str, Any]:
    """Aplica el plan. Cada alta usa el escritor canónico e idempotente."""
    plan = planificar_importacion_costos_waimao(ruta_xlsx)
    aplicados = []
    for item in plan["candidatos"]:
        origen = {
            "fuente": ORIGEN,
            "source_sha256": plan["source_sha256"],
            "sheet": HOJA,
            "fila_sheet": item["fila_sheet"],
            "campo_costo": "COSTOINICIAL",
            "tracking": item["tracking"],
            "empresa_sheet": item["empresa_sheet"],
        }
        try:
            resultado = registrar_snapshot_cotizacion(
                solicitud_id=int(item["solicitud_id"]),
                coti_id=item.get("coti_id"),
                courier=item["courier"],
                servicio_courier=item.get("servicio_courier"),
                moneda_courier="ARS",
                tipo_cambio_ars=1,
                costo_courier_estimado=item["costo_inicial_ars"],
                precio_cliente_inicial_ars=item["precio_tauro_ars"],
                margen_tauro_protegido_ars=item["margen_tauro_ars"],
                peso_real_cotizado_kg=item.get("peso_kg"),
                peso_facturable_cotizado_kg=item.get("peso_kg"),
                bultos=item.get("bultos") or [],
                origen_calculo=origen,
                aceptado_at=item.get("created_at"),
                actor=actor,
            )
        except ConciliacionCourierError as exc:
            raise ImportacionCostoWaimaoError(
                f"No se pudo importar {item['tracking']}: {exc}"
            ) from exc
        aplicados.append({
            "solicitud_id": int(item["solicitud_id"]),
            "tracking": item["tracking"],
            "snapshot_id": int(resultado["id"]),
            "duplicado": bool(resultado["duplicado"]),
        })
    plan["aplicados"] = aplicados
    return plan
