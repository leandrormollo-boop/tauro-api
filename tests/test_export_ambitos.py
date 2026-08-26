"""El backup conserva envíos y libros contables separados y reconciliados."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
import inspect

import pytest
from openpyxl import load_workbook

from servicios import catalogo, cuenta_corriente, export_cliente, solicitudes_guia


def _envio(numero, courier, origen, destino):
    return {
        "id": numero, "created_at": datetime(2026, 8, numero),
        "estado": "GUIA_LISTA", "courier": courier,
        "remitente_pais": origen, "destino_pais": destino,
        "tracking": str(numero), "producto_alias": "CARGA", "cantidad": 1,
        "dest_nombre": "Persona", "dest_ciudad": "Ciudad",
        "peso_kg": 1, "valor_declarado_usd": 10,
        "precio_tauro_ars": 1000, "precio_tauro_usd": 1,
        "observaciones": "",
    }


def _ledger(facturado, pendiente, debe, haber, saldo):
    return {
        "facturado_ars": Decimal(facturado),
        "pendiente_facturacion_ars": Decimal(pendiente),
        "debe_ars": Decimal(debe),
        "haber_ars": Decimal(haber),
        "saldo_ars": Decimal(saldo),
    }


def _resumen():
    return {
        "nacional": _ledger("100.10", "0.00", "100.10", "40.05", "60.05"),
        "internacional": _ledger(
            "150.10", "50.10", "200.20", "100.10", "100.10"
        ),
        "consolidado": _ledger(
            "250.20", "75.35", "325.55", "170.30", "155.25"
        ),
        "credito_sin_imputar_ars": Decimal("30.15"),
        "cargos_sin_clasificar_ars": Decimal("25.25"),
    }


def _movimiento(tipo, ambito, debe="0.00", haber="0.00", numero=1, estado="ACTIVO"):
    return {
        "fecha": f"{numero:02d}/08/2026", "tipo": tipo, "ambito": ambito,
        "concepto": f"Movimiento {numero}", "referencia": f"REF-{numero}",
        "estado": estado, "debe_ars": Decimal(debe),
        "haber_ars": Decimal(haber), "archivo_url": "",
    }


MOVIMIENTOS_CONSOLIDADOS = [
    _movimiento("FC", "NACIONAL", debe="100.10", numero=1),
    _movimiento("PENDIENTE_FACTURA", "INTERNACIONAL", debe="200.20", numero=2),
    _movimiento("PAGO", "NACIONAL", haber="40.05", numero=3, estado="APROBADO"),
    _movimiento("PAGO", "INTERNACIONAL", haber="100.10", numero=4, estado="APROBADO"),
    _movimiento("PAGO", "SIN_IMPUTAR", haber="30.15", numero=5, estado="APROBADO"),
    _movimiento("PENDIENTE_FACTURA", "SIN_CLASIFICAR", debe="25.25", numero=6),
    _movimiento("PAGO_PENDIENTE", "SIN_IMPUTAR", numero=7, estado="PENDIENTE"),
]


def _instalar_fakes(monkeypatch, *, resumen=None, movimientos=None):
    llamadas = []
    monkeypatch.setattr(
        solicitudes_guia, "listar_solicitudes_cliente",
        lambda *_args, **_kwargs: [
            _envio(1, "DHL", "AR", "AR"),
            _envio(2, "OCA", "US", "AR"),
        ],
    )
    monkeypatch.setattr(catalogo, "get_productos", lambda *_a, **_k: [])
    monkeypatch.setattr(
        cuenta_corriente, "resumen_cuenta_por_ambito",
        lambda _cliente: resumen or _resumen(),
    )

    por_ambito = movimientos or {
        "consolidado": MOVIMIENTOS_CONSOLIDADOS,
        "nacional": [m for m in MOVIMIENTOS_CONSOLIDADOS if m["ambito"] == "NACIONAL"],
        "internacional": [m for m in MOVIMIENTOS_CONSOLIDADOS if m["ambito"] == "INTERNACIONAL"],
    }

    def paginados(cliente, ambito, tipo, pagina, page_size):
        llamadas.append((cliente, ambito, tipo, pagina, page_size))
        todos = por_ambito[ambito]
        # Consolidado ocupa dos páginas en el fake para probar que el export
        # avanza hasta completar el historial sin pedir una lista ilimitada.
        corte = 3 if ambito == "consolidado" else len(todos) or 1
        paginas = [todos[i:i + corte] for i in range(0, len(todos), corte)] or [[]]
        return {
            "items": paginas[pagina - 1], "pagina_actual": pagina,
            "total_paginas": len(paginas), "total_resultados": len(todos),
            "pagina_desde": 1, "pagina_hasta": len(paginas[pagina - 1]),
            "paginas_visibles": list(range(1, len(paginas) + 1)),
        }

    monkeypatch.setattr(cuenta_corriente, "movimientos_cuenta_paginados", paginados)
    return llamadas


def _cierre(ws, etiqueta):
    for fila in ws.iter_rows(min_row=2):
        if fila[5].value == etiqueta:
            valor = fila[6].value if fila[6].value not in (None, "") else fila[7].value
            return Decimal(str(valor))
    raise AssertionError(f"No se encontró {etiqueta!r} en {ws.title}")


def test_excel_crea_envios_y_cuentas_separadas_reconciliadas(monkeypatch):
    llamadas = _instalar_fakes(monkeypatch)

    libro = load_workbook(BytesIO(export_cliente.generar_excel_cliente(" test ")))

    assert "Envíos" not in libro.sheetnames
    assert "Envios_Nacionales" in libro.sheetnames
    assert "Envios_Internacionales" in libro.sheetnames
    assert libro["Envios_Nacionales"].max_row == 2
    assert libro["Envios_Internacionales"].max_row == 2
    for titulo in (
        "Cuenta_Consolidada", "Cuenta_Nacional", "Cuenta_Internacional",
        "Cuenta_Sin_imputar", "Cuenta_Sin_clasificar",
    ):
        assert titulo in libro.sheetnames

    consolidada = libro["Cuenta_Consolidada"]
    nacional = libro["Cuenta_Nacional"]
    internacional = libro["Cuenta_Internacional"]
    assert _cierre(consolidada, "TOTAL DEBE") == Decimal("325.55")
    assert _cierre(consolidada, "TOTAL HABER") == Decimal("170.30")
    assert _cierre(consolidada, "SALDO (DEBE - HABER)") == Decimal("155.25")
    assert _cierre(nacional, "TOTAL HABER") == Decimal("40.05")
    assert _cierre(internacional, "TOTAL HABER") == Decimal("100.10")
    assert _cierre(libro["Cuenta_Sin_imputar"], "TOTAL HABER") == Decimal("30.15")
    assert _cierre(libro["Cuenta_Sin_clasificar"], "TOTAL DEBE") == Decimal("25.25")

    # Se solicitaron páginas acotadas y la segunda página del consolidado.
    assert llamadas
    assert all(llamada[0] == "TEST" and llamada[2] == "todos" for llamada in llamadas)
    assert all(llamada[4] == 100 for llamada in llamadas)
    assert ("TEST", "consolidado", "todos", 2, 100) in llamadas
    assert not any(llamada[3] > 3 for llamada in llamadas)


def test_credito_historico_no_se_autoimputa_a_nacional_o_internacional(monkeypatch):
    _instalar_fakes(monkeypatch)
    libro = load_workbook(BytesIO(export_cliente.generar_excel_cliente("TEST")))

    assert _cierre(libro["Cuenta_Nacional"], "TOTAL HABER") == Decimal("40.05")
    assert _cierre(libro["Cuenta_Internacional"], "TOTAL HABER") == Decimal("100.10")
    assert _cierre(libro["Cuenta_Sin_imputar"], "TOTAL HABER") == Decimal("30.15")
    assert (
        _cierre(libro["Cuenta_Nacional"], "TOTAL HABER")
        + _cierre(libro["Cuenta_Internacional"], "TOTAL HABER")
        + _cierre(libro["Cuenta_Sin_imputar"], "TOTAL HABER")
    ) == _cierre(libro["Cuenta_Consolidada"], "TOTAL HABER")


def test_hojas_residuales_no_aparecen_si_no_hay_partidas(monkeypatch):
    resumen = {
        "nacional": _ledger("100.10", "0", "100.10", "40.05", "60.05"),
        "internacional": _ledger("0", "200.20", "200.20", "100.10", "100.10"),
        "consolidado": _ledger("100.10", "200.20", "300.30", "140.15", "160.15"),
        "credito_sin_imputar_ars": Decimal("0"),
        "cargos_sin_clasificar_ars": Decimal("0"),
    }
    movimientos = {
        "consolidado": [m for m in MOVIMIENTOS_CONSOLIDADOS if m["ambito"] in {"NACIONAL", "INTERNACIONAL"}],
        "nacional": [m for m in MOVIMIENTOS_CONSOLIDADOS if m["ambito"] == "NACIONAL"],
        "internacional": [m for m in MOVIMIENTOS_CONSOLIDADOS if m["ambito"] == "INTERNACIONAL"],
    }
    _instalar_fakes(monkeypatch, resumen=resumen, movimientos=movimientos)

    libro = load_workbook(BytesIO(export_cliente.generar_excel_cliente("TEST")))

    assert "Cuenta_Sin_imputar" not in libro.sheetnames
    assert "Cuenta_Sin_clasificar" not in libro.sheetnames


def test_export_frena_si_los_libros_no_reconcilian(monkeypatch):
    resumen = _resumen()
    resumen["consolidado"] = {**resumen["consolidado"], "saldo_ars": Decimal("999")}
    _instalar_fakes(monkeypatch, resumen=resumen)

    with pytest.raises(RuntimeError, match="saldo consolidado no cierra"):
        export_cliente.generar_excel_cliente("TEST")


def test_dinero_de_cuenta_no_pasa_por_float():
    fuente = inspect.getsource(export_cliente._fila_movimiento_cuenta)
    assert "Decimal(str(" in fuente
    assert "float(" not in fuente
